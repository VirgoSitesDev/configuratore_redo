from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import os
import json
from database import DatabaseManager
from datetime import datetime
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
import math
from admin import admin_bp
from flask_mail import Mail, Message
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

app = Flask(__name__)
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', '587'))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() in ['true', 'on', '1']
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER')

mail = Mail(app)
CORS(app)
app.register_blueprint(admin_bp)
app.secret_key = os.environ.get('SECRET_KEY', 'ju16i_8nf&+o766zi79z0_dkk8l$2g!no7&dzfcrhcw_%&_4w4')

db = DatabaseManager()
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get_categorie')
def get_categorie():
    try:
        categorie = db.get_categorie()
        return jsonify(categorie)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_profili/<categoria>')
def get_profili(categoria):
    try:
        profili = db.get_profili_by_categoria(categoria)

        tutti_strip_ids = set()
        for profilo in profili:
            strip_ids = profilo.get('stripLedCompatibili', [])
            tutti_strip_ids.update(strip_ids)

        strip_led_map = {}
        if tutti_strip_ids:
            strip_ids_list = list(tutti_strip_ids)

            strip_data_response = db.supabase.table('strip_led')\
                .select('id, nome_commerciale')\
                .in_('id', strip_ids_list)\
                .execute()

            if strip_data_response.data:
                for strip in strip_data_response.data:
                    strip_led_map[strip['id']] = {
                        'id': strip['id'],
                        'nomeCommerciale': strip.get('nome_commerciale', '')
                    }

        for profilo in profili:
            strip_compatibili_info = []
            for strip_id in profilo.get('stripLedCompatibili', []):
                if strip_id in strip_led_map:
                    strip_compatibili_info.append(strip_led_map[strip_id])
            profilo['stripLedCompatibiliInfo'] = strip_compatibili_info
        
        return jsonify(profili)
    except Exception as e:
        logging.error(f"Errore in get_profili: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_opzioni_profilo/<profilo_id>')
def get_opzioni_profilo(profilo_id):
    try:
        profilo_data = db.supabase.table('profili').select('*').eq('id', profilo_id).single().execute().data
        
        if not profilo_data:
            return jsonify({'tipologie': []})
            
        tipologie = db.supabase.table('profili_tipologie').select('tipologia').eq('profilo_id', profilo_id).execute().data
        
        return jsonify({
            'tipologie': [t['tipologia'] for t in tipologie]
        })
    except Exception as e:
        return jsonify({'tipologie': []})

@app.route('/get_opzioni_tensione/<profilo_id>')
@app.route('/get_opzioni_tensione/<profilo_id>/<tipologia_strip>')
def get_opzioni_tensione(profilo_id, tipologia_strip=None):
    try:
        if profilo_id == 'ESTERNI':
            voltaggi_disponibili = ['24V', '48V', '220V']
            if tipologia_strip == 'SPECIAL':
                voltaggi_disponibili = ['24V']
            return jsonify({'success': True, 'voltaggi': voltaggi_disponibili})
        
        strip_compatibili = db.supabase.table('profili_strip_compatibili').select('strip_id').eq('profilo_id', profilo_id).execute().data
        strip_ids = [s['strip_id'] for s in strip_compatibili]
        
        if not strip_ids:
            return jsonify({'success': True, 'voltaggi': []})
        
        query = db.supabase.table('strip_led').select('tensione')
        
        if tipologia_strip:
            query = query.eq('tipo', tipologia_strip)
        
        strips = query.in_('id', strip_ids).execute().data
        
        voltaggi_disponibili = list(set([s['tensione'] for s in strips]))
        
        return jsonify({'success': True, 'voltaggi': voltaggi_disponibili})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_opzioni_ip/<profilo_id>/<tensione>')
@app.route('/get_opzioni_ip/<profilo_id>/<tensione>/<tipologia_strip>')
def get_opzioni_ip(profilo_id, tensione, tipologia_strip=None):
    try:
        strip_compatibili = db.supabase.table('profili_strip_compatibili').select('strip_id').eq('profilo_id', profilo_id).execute().data
        strip_ids = [s['strip_id'] for s in strip_compatibili]
        
        if not strip_ids:
            return jsonify({'success': True, 'ip': []})
        
        query = db.supabase.table('strip_led').select('ip').eq('tensione', tensione)
        
        if tipologia_strip:
            query = query.eq('tipo', tipologia_strip)
        
        strips = query.in_('id', strip_ids).execute().data
        
        ip_disponibili = list(set([s['ip'] for s in strips]))
        
        return jsonify({'success': True, 'ip': ip_disponibili})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_opzioni_temperatura_iniziale/<profilo_id>/<tensione>/<ip>')
@app.route('/get_opzioni_temperatura_iniziale/<profilo_id>/<tensione>/<ip>/<tipologia_strip>')
def get_opzioni_temperatura_iniziale(profilo_id, tensione, ip, tipologia_strip=None):
    try:
        strip_compatibili = db.supabase.table('profili_strip_compatibili').select('strip_id').eq('profilo_id', profilo_id).execute().data
        strip_ids = [s['strip_id'] for s in strip_compatibili]
        
        if not strip_ids:
            return jsonify({'success': True, 'temperature': []})
        
        query = db.supabase.table('strip_led').select('id').eq('tensione', tensione).eq('ip', ip)
        
        if tipologia_strip:
            query = query.eq('tipo', tipologia_strip)
        
        strips = query.in_('id', strip_ids).execute().data
        
        temperature_disponibili = set()
        for strip in strips:
            temps = db.supabase.table('strip_temperature').select('temperatura').eq('strip_id', strip['id']).execute().data
            for t in temps:
                temperature_disponibili.add(t['temperatura'])
        
        return jsonify({'success': True, 'temperature': list(temperature_disponibili)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_temperature_colore/<profilo_id>/<tipologia>/<tensione>/<ip>')
def get_temperature_colore(profilo_id, tipologia, tensione, ip):
    try:
        strips = db.get_strip_led_filtrate(profilo_id, tensione, ip, None, None, tipologia)
        
        temperature = set()
        for strip in strips:
            for temp in strip.get('temperaturaColoreDisponibili', []):
                temperature.add(temp)
        
        temperature_list = list(temperature)
        temperature_list.sort()
        
        return jsonify({'success': True, 'temperature': temperature_list})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_opzioni_potenza/<profilo_id>/<tensione>/<ip>/<temperatura>')
@app.route('/get_opzioni_potenza/<profilo_id>/<tensione>/<ip>/<temperatura>/<tipologia_strip>')
def get_opzioni_potenza(profilo_id, tensione, ip, temperatura, tipologia_strip=None):
    try:
        if profilo_id == 'ESTERNI':
            strips = db.get_all_strip_led_filtrate(tensione, ip, temperatura, None, tipologia_strip)
        else:
            strips = db.get_strip_led_filtrate(profilo_id, tensione, ip, temperatura, None, tipologia_strip)
        
        tutte_potenze_disponibili = set()
        for strip in strips:
            potenze_strip = strip.get('potenzeDisponibili', [])
            tutte_potenze_disponibili.update(potenze_strip)
        
        if not tutte_potenze_disponibili:
            return jsonify({'success': False, 'message': 'Nessuna potenza disponibile per i parametri selezionati'})
        
        potenze_complete = []
        for potenza in tutte_potenze_disponibili:
            potenze_complete.append({
                'id': potenza,
                'nome': potenza,
                'codice': '',
                'specifiche': ''
            })

        return jsonify({'success': True, 'potenze': potenze_complete})
        
    except Exception as e:
        logging.error(f"ERRORE GRAVE in get_opzioni_potenza: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_strip_led_filtrate/<profilo_id>/<tensione>/<ip>/<temperatura>/<potenza>')
@app.route('/get_strip_led_filtrate/<profilo_id>/<tensione>/<ip>/<temperatura>/<potenza>/<tipologia>')
def get_strip_led_filtrate(profilo_id, tensione, ip, temperatura, potenza, tipologia=None):
    try:
        if potenza:
            potenza = potenza.replace('-', ' ').replace('_', '/')

        if profilo_id == 'ESTERNI':
            strips = db.get_all_strip_led_filtrate(tensione, ip, temperatura, potenza, tipologia)
        else:
            strips = db.get_strip_led_filtrate(profilo_id, tensione, ip, temperatura, potenza, tipologia)
        
        return jsonify({'success': True, 'strip_led': strips})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_opzioni_alimentatore/<tipo_alimentazione>/<tensione_strip>')
@app.route('/get_opzioni_alimentatore/<tipo_alimentazione>/<tensione_strip>/<int:potenza_consigliata>')
def get_opzioni_alimentatore(tipo_alimentazione, tensione_strip, potenza_consigliata=None):
    try:
        alimentatori = db.get_alimentatori_by_tipo(tipo_alimentazione, tensione_strip)
        
        if potenza_consigliata:
            alimentatori_filtrati = []
            for alim in alimentatori:
                potenze_adatte = [p for p in alim.get('potenze', []) if p >= potenza_consigliata]
                if potenze_adatte:
                    alim['potenza_consigliata'] = min(potenze_adatte)
                    alimentatori_filtrati.append(alim)
        else:
            alimentatori_filtrati = alimentatori
        
        dettagli_alimentatori = {alim['id']: alim for alim in alimentatori}
        
        return jsonify({
            'success': True,
            'alimentatori': alimentatori_filtrati,
            'dettagliAlimentatori': dettagli_alimentatori
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_potenze_alimentatore/<alimentatore_id>')
def get_potenze_alimentatore(alimentatore_id):
    try:
        potenze = db.get_potenze_alimentatore(alimentatore_id)
        return jsonify({'success': True, 'potenze': potenze})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_dettagli_alimentatore/<alimentatore_id>')
def get_dettagli_alimentatore(alimentatore_id):
    try:
        alimentatore = db.get_dettagli_alimentatore(alimentatore_id)
        
        if alimentatore:
            return jsonify({'success': True, 'alimentatore': alimentatore})
        else:
            return jsonify({'success': False, 'message': 'Alimentatore non trovato'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_opzioni_dimmerazione/<strip_id>')
def get_opzioni_dimmerazione(strip_id):
    try:
        result = db.get_dimmer_compatibili(strip_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_dimmer_compatibili/<strip_id>')
def get_dimmer_compatibili(strip_id):
    try:
        result = db.get_dimmer_compatibili(strip_id)
        dimmer_compatibili = [d for d in result.get('opzioni', []) if d != 'NESSUN_DIMMER']
        return jsonify({'success': True, 'dimmer_compatibili': dimmer_compatibili})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_finiture/<profilo_id>')
def get_finiture(profilo_id):
    try:
        finiture = db.supabase.table('profili_finiture').select('finitura').eq('profilo_id', profilo_id).execute().data
        finiture_list = [f['finitura'] for f in finiture]
        
        mappatura_finiture = {
            'ALLUMINIO_ANODIZZATO': 'Alluminio anodizzato',
            'BIANCO': 'Bianco',
            'NERO': 'Nero',
            'ALLUMINIO': 'Alluminio'
        }
        
        finiture_formattate = [
            {
                'id': finitura,
                'nome': mappatura_finiture.get(finitura, finitura)
            }
            for finitura in finiture_list
        ]
        
        return jsonify({'success': True, 'finiture': finiture_formattate})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/calcola_potenza_alimentatore', methods=['POST'])
def calcola_potenza_alimentatore():
    data = request.json
    potenza_per_metro = data.get('potenzaPerMetro', 0)
    lunghezza_metri = data.get('lunghezzaMetri', 0)
    potenza_totale = potenza_per_metro * lunghezza_metri * 1.2
    potenze_standard = [30, 60, 100, 150, 200, 320]
    potenza_consigliata = next((p for p in potenze_standard if p >= potenza_totale), potenze_standard[-1])
    
    return jsonify({
        'success': True,
        'potenzaTotale': round(potenza_totale, 2),
        'potenzaConsigliata': potenza_consigliata
    })

@app.route('/calcola_potenza_consigliata', methods=['POST'])
def calcola_potenza_consigliata():
    try:
        data = request.json
        
        potenza_strip = float(data.get('potenzaStrip', 0))
        lunghezza = float(data.get('lunghezza', 0))
        alimentazione_doppia = data.get('alimentazioneDoppia', False)
        
        potenza_totale = (potenza_strip * lunghezza) / 1000
        
        if alimentazione_doppia:
            potenza_totale = potenza_totale / 2
        
        potenza_consigliata = potenza_totale * 1.2
        potenza_consigliata = int((potenza_consigliata + 4) / 5) * 5
        
        return jsonify({
            'success': True,
            'potenzaConsigliata': potenza_consigliata,
            'potenzaTotale': potenza_totale
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_strip_compatibile_standalone', methods=['POST'])
def get_strip_compatibile_standalone():
    """Trova una strip reale dal database invece di generare ID fittizi"""
    try:
        data = request.json
        
        tipologia = data.get('tipologia')
        tensione = data.get('tensione')
        ip = data.get('ip')
        temperatura = data.get('temperatura')
        potenza = data.get('potenza')
        special = data.get('special')
        query = db.supabase.table('strip_led').select('*')
        query = query.eq('tensione', tensione).eq('ip', ip)

        if tipologia == 'SPECIAL' and special:
            if special == 'XMAGIS':
                query = query.or_('nome_commerciale.ilike.%XMAGIS%,id.ilike.%XMAGIS%,nome_commerciale.ilike.%MG13X12%,nome_commerciale.ilike.%MG12X17%')
            elif special == 'XFLEX':
                query = query.or_('nome_commerciale.ilike.%XFLEX%,id.ilike.%XFLEX%')
            elif special == 'XSNAKE':
                query = query.or_('nome_commerciale.ilike.%XSNAKE%,id.ilike.%XSNAKE%,id.ilike.%SNK%')
            elif special == 'ZIG_ZAG':
                query = query.or_('nome_commerciale.ilike.%ZIGZAG%,id.ilike.%ZIGZAG%,nome_commerciale.ilike.%ZIG_ZAG%')
        elif tipologia and tipologia != 'None':
            query = query.eq('tipo', tipologia)
        
        strips = query.execute().data
        
        if not strips:
            return jsonify({
                'success': False, 
                'message': 'Nessuna strip trovata per i parametri specificati'
            })

        if temperatura:
            strip_ids = [s['id'] for s in strips]
            temp_check = db.supabase.table('strip_temperature')\
                .select('strip_id')\
                .eq('temperatura', temperatura)\
                .in_('strip_id', strip_ids)\
                .execute().data
            
            strip_ids_con_temp = [t['strip_id'] for t in temp_check]
            strips = [s for s in strips if s['id'] in strip_ids_con_temp]

        if potenza:
            strip_ids = [s['id'] for s in strips]
            potenza_check = db.supabase.table('strip_potenze')\
                .select('strip_id')\
                .eq('potenza', potenza)\
                .in_('strip_id', strip_ids)\
                .execute().data
            
            strip_ids_con_potenza = [p['strip_id'] for p in potenza_check]
            strips = [s for s in strips if s['id'] in strip_ids_con_potenza]
        
        if not strips:
            return jsonify({
                'success': False, 
                'message': 'Nessuna strip trovata dopo tutti i filtri'
            })

        strip_scelta = strips[0]
        logging.info(f"Strip scelta: {strip_scelta['id']} - {strip_scelta.get('nome_commerciale', '')}")
        
        return jsonify({
            'success': True,
            'strip_led': {
                'id': strip_scelta['id'],
                'nomeCommerciale': strip_scelta.get('nome_commerciale', ''),
                'nome': strip_scelta.get('nome', ''),
                'tensione': strip_scelta.get('tensione', ''),
                'ip': strip_scelta.get('ip', ''),
                'tipo': strip_scelta.get('tipo', '')
            }
        })
        
    except Exception as e:
        logging.error(f"Errore in get_strip_compatibile_standalone: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/calcola_lunghezze', methods=['POST'])
def calcola_lunghezze():
    data = request.json
    dim_richiesta = data.get('lunghezzaRichiesta', 0)
    strip_id = data.get('stripLedSelezionata')
    potenza_selezionata = data.get('potenzaSelezionata')
    lunghezze_multiple = data.get('lunghezzeMultiple', {})
    forma_taglio = data.get('formaDiTaglioSelezionata', 'DRITTO_SEMPLICE')
    
    taglio_minimo = 1
    spazio_produzione = 5

    if strip_id and strip_id != 'NO_STRIP' and potenza_selezionata:
        try:
            strip_data_result = db.supabase.table('strip_led').select('*').eq('id', strip_id).execute()
            
            if strip_data_result.data and len(strip_data_result.data) > 0:
                strip_info = strip_data_result.data[0]
                tagli_minimi = strip_info.get('taglio_minimo', [])
                potenze_data = db.supabase.table('strip_potenze').select('*').eq('strip_id', strip_id).order('indice').execute()
                if potenze_data.data:
                    indice_potenza = -1
                    for record in potenze_data.data:
                        if record.get('potenza') == potenza_selezionata:
                            indice_potenza = record.get('indice', -1)
                            break

                    if indice_potenza >= 0 and indice_potenza < len(tagli_minimi):
                        taglio_minimo_str = tagli_minimi[indice_potenza]
                        import re
                        match = re.search(r'(\d+(?:[.,]\d+)?)', taglio_minimo_str)
                        if match:
                            taglio_minimo_val = match.group(1).replace(',', '.')
                            try:
                                taglio_minimo = float(taglio_minimo_val)
                            except ValueError:
                                logging.warning(f"Impossibile convertire taglio_minimo: {taglio_minimo_val}")
                                pass
            else:
                logging.warning(f"Strip non trovata nel database: {strip_id}")
                
        except Exception as e:
            logging.error(f"Errore nella ricerca della strip {strip_id}: {str(e)}")

    def calcola_proposte_singole(lunghezza):
        if lunghezza > 0:
            proposta1 = ((lunghezza - 5) // taglio_minimo * taglio_minimo) + 5
            proposta2 = ((((lunghezza - 5) + taglio_minimo - 0.01) // taglio_minimo) * taglio_minimo) + 5

            if proposta2 <= proposta1:
                proposta2 = (proposta1 + taglio_minimo)
        else:
            proposta1 = 0
            proposta2 = 0

        if data.get('formaDiTaglioSelezionata') is None:
            proposta1 -= 5
            proposta2 -= 5
        return proposta1, proposta2

    def calcola_spazio_buio_lato(valore_originale, proposta1, proposta2):
        if valore_originale == proposta1 or valore_originale == proposta2:
            return 0

        if proposta1 < valore_originale < proposta2:
            return valore_originale - proposta1

        if valore_originale < proposta1:
            return 0

        if valore_originale > proposta2:
            return valore_originale - proposta2
        
        return 0

    if forma_taglio != 'DRITTO_SEMPLICE' and lunghezze_multiple:
        proposte_per_lato = {}
        
        for lato, lunghezza in lunghezze_multiple.items():
            if lunghezza and lunghezza > 0:
                prop1, prop2 = calcola_proposte_singole(lunghezza)
                proposte_per_lato[lato] = {
                    'originale': lunghezza,
                    'proposta1': prop1,
                    'proposta2': prop2
                }

        import itertools
        
        lati_ordinati = sorted(proposte_per_lato.keys())
        combinazioni = []
        opzioni_per_lato = []
        for lato in lati_ordinati:
            opzioni = []
            props = proposte_per_lato[lato]
            
            if props['proposta1'] != props['originale']:
                opzioni.append(('proposta1', props['proposta1']))
            
            if props['proposta2'] != props['originale'] and props['proposta2'] != props['proposta1']:
                opzioni.append(('proposta2', props['proposta2']))
            
            opzioni.append(('originale', props['originale']))
            opzioni_per_lato.append(opzioni)

        for combinazione in itertools.product(*opzioni_per_lato):
            combo_dict = {}
            combo_label_parts = []
            lunghezza_totale = 0
            spazio_buio_totale = 0
            
            for i, (tipo, valore) in enumerate(combinazione):
                lato = lati_ordinati[i]
                combo_dict[lato] = valore
                lunghezza_totale += valore

                props = proposte_per_lato[lato]
                spazio_buio_lato = calcola_spazio_buio_lato(
                    props['originale'], 
                    props['proposta1'], 
                    props['proposta2']
                ) if tipo == 'originale' else 0
                
                spazio_buio_totale += spazio_buio_lato

                if tipo == 'originale':
                    combo_label_parts.append(f"Orig.")
                elif tipo == 'proposta1':
                    combo_label_parts.append(f"Prop.1")
                else:
                    combo_label_parts.append(f"Prop.2")
            
            combinazioni.append({
                'lunghezze': combo_dict,
                'lunghezza_totale': lunghezza_totale,
                'label': " + ".join(combo_label_parts),
                'ha_spazio_buio': spazio_buio_totale > 0,
                'spazio_buio_totale': spazio_buio_totale,
                'dettaglio': combinazione
            })

        combinazioni.sort(key=lambda x: (x['ha_spazio_buio'], x['spazio_buio_totale'], x['lunghezza_totale']))
        
        return jsonify({
            'success': True,
            'tipo': 'combinazioni',
            'spazioProduzione': spazio_produzione,
            'proposte_per_lato': proposte_per_lato,
            'combinazioni': combinazioni[:27]
        })
    
    else:
        if dim_richiesta > 0:
            proposta1, proposta2 = calcola_proposte_singole(dim_richiesta)
        else:
            proposta1 = 0
            proposta2 = 0
        
        return jsonify({
            'success': True,
            'tipo': 'semplice',
            'spazioProduzione': spazio_produzione,
            'proposte': {
                'proposta1': proposta1,
                'proposta2': proposta2
            }
        })

def ottimizza_quantita_profilo(lunghezza_richiesta, lunghezze_disponibili):
    """
    Trova la combinazione ottimale di lunghezze standard per minimizzare lo spreco
    usando programmazione dinamica per esplorare tutte le combinazioni
    """
    if not lunghezze_disponibili or lunghezza_richiesta <= 0:
        return [{'lunghezza': 3000, 'quantita': 1}]

    lunghezze = sorted(list(set(lunghezze_disponibili)))
    limite_massimo = lunghezza_richiesta + max(lunghezze)
    
    migliore_combinazione = None
    minimo_spreco = float('inf')
    minimo_pezzi = float('inf')

    def genera_combinazioni(lunghezza_target, combinazione_corrente, lunghezze_usate):
        nonlocal migliore_combinazione, minimo_spreco, minimo_pezzi

        lunghezza_totale = sum(lunghezze_usate)

        if lunghezza_totale >= lunghezza_target:
            spreco = lunghezza_totale - lunghezza_target
            pezzi_totali = sum(combinazione_corrente.values())

            if (spreco < minimo_spreco) or (spreco == minimo_spreco and pezzi_totali < minimo_pezzi):
                minimo_spreco = spreco
                minimo_pezzi = pezzi_totali
                migliore_combinazione = [
                    {'lunghezza': lunghezza, 'quantita': quantita}
                    for lunghezza, quantita in combinazione_corrente.items()
                    if quantita > 0
                ]
            return

        if lunghezza_totale > limite_massimo:
            return

        for lunghezza in lunghezze:
            max_pezzi_per_lunghezza = min(10, (limite_massimo // lunghezza) + 1)
            
            if combinazione_corrente.get(lunghezza, 0) < max_pezzi_per_lunghezza:
                nuova_combinazione = combinazione_corrente.copy()
                nuova_combinazione[lunghezza] = nuova_combinazione.get(lunghezza, 0) + 1
                nuove_lunghezze = lunghezze_usate + [lunghezza]
                
                genera_combinazioni(lunghezza_target, nuova_combinazione, nuove_lunghezze)

    genera_combinazioni(lunghezza_richiesta, {}, [])

    if migliore_combinazione is None:
        lunghezza_max = max(lunghezze)
        quantita_necessaria = math.ceil(lunghezza_richiesta / lunghezza_max)
        migliore_combinazione = [{'lunghezza': lunghezza_max, 'quantita': quantita_necessaria}]

    migliore_combinazione.sort(key=lambda x: x['lunghezza'], reverse=True)
    
    return migliore_combinazione

@app.route('/finalizza_configurazione', methods=['POST'])
def finalizza_configurazione():
    configurazione = request.json

    def calcola_codici_prodotto():
        codici = {
            'profilo': '',
            'stripLed': '',
            'alimentatore': '',
            'dimmer': ''
        }

        if configurazione.get('codiceProfilo'):
            codici['profilo'] = configurazione['codiceProfilo']
        elif configurazione.get('profiloSelezionato'):
            codici['profilo'] = configurazione['profiloSelezionato'].replace('_', '/')
            
        if configurazione.get('stripLedSelezionata'):
            codici['stripLed'] = configurazione['stripLedSelezionata']
            
        if configurazione.get('codiceAlimentatore'):
            codici['alimentatore'] = configurazione['codiceAlimentatore']
            
        if configurazione.get('codiceDimmer'):
            codici['dimmer'] = configurazione['codiceDimmer']
        elif configurazione.get('dimmerSelezionato') == 'NESSUN_DIMMER':
            codici['dimmer'] = ''
        elif configurazione.get('dimmerCodice'):
            codici['dimmer'] = configurazione['dimmerCodice']
            
        return codici

    tuttiCodici = calcola_codici_prodotto()
    
    potenza_per_metro = 0
    if 'potenzaSelezionata' in configurazione and configurazione['potenzaSelezionata']:
        potenza_str = configurazione['potenzaSelezionata'].split('W/m')[0]
        try:
            potenza_per_metro = float(potenza_str)
        except ValueError:
            potenza_per_metro = 0
    
    lunghezza_in_metri = 0
    if 'lunghezzaRichiesta' in configurazione and configurazione['lunghezzaRichiesta']:
        try:
            lunghezza_in_metri = float(configurazione['lunghezzaRichiesta']) / 1000
        except ValueError:
            lunghezza_in_metri = 0
    
    potenza_totale = potenza_per_metro * lunghezza_in_metri

    quantita_profilo = 1
    quantita_strip_led = 1
    lunghezza_massima_profilo = 3000
    lunghezza_massima_strip = 5000
    combinazione_ottimale = [{'lunghezza': 3000, 'quantita': 1}]

    lunghezza_totale = 0
    if 'lunghezzeMultiple' in configurazione and configurazione['lunghezzeMultiple']:
        lunghezza_totale = sum(v for v in configurazione['lunghezzeMultiple'].values() if v and v > 0)
    elif 'lunghezzaRichiesta' in configurazione and configurazione['lunghezzaRichiesta']:
        lunghezza_totale = float(configurazione['lunghezzaRichiesta'])

    if 'profiloSelezionato' in configurazione and configurazione['profiloSelezionato']:
        try:
            lunghezze_profilo = db.supabase.table('profili_lunghezze')\
                .select('lunghezza')\
                .eq('profilo_id', configurazione['profiloSelezionato'])\
                .execute().data
            
            if lunghezze_profilo and lunghezza_totale > 0:
                lunghezze_list = [l['lunghezza'] for l in lunghezze_profilo]
                lunghezza_massima_profilo = max(lunghezze_list)
                combinazione_ottimale = ottimizza_quantita_profilo(lunghezza_totale, lunghezze_list)
                quantita_profilo = sum(item['quantita'] for item in combinazione_ottimale)
                
            elif lunghezze_profilo:
                lunghezze_list = [l['lunghezza'] for l in lunghezze_profilo]
                lunghezza_massima_profilo = max(lunghezze_list)
                quantita_profilo = 1
                combinazione_ottimale = [{'lunghezza': lunghezza_massima_profilo, 'quantita': 1}]
                
        except Exception as e:
            logging.error(f"Errore nel calcolo ottimizzato quantità profilo: {str(e)}")
            if lunghezza_totale > 0:
                quantita_profilo = math.ceil(lunghezza_totale / lunghezza_massima_profilo)
            combinazione_ottimale = [{'lunghezza': lunghezza_massima_profilo, 'quantita': quantita_profilo}]

    if 'stripLedSelezionata' in configurazione and configurazione['stripLedSelezionata'] and configurazione['stripLedSelezionata'] not in ['NO_STRIP', 'senza_strip']:
        try:
            strip_data = db.supabase.table('strip_led')\
                .select('lunghezza')\
                .eq('id', configurazione['stripLedSelezionata'])\
                .single()\
                .execute()
            
            if strip_data.data and strip_data.data.get('lunghezza'):
                lunghezza_massima_strip = strip_data.data['lunghezza']
                logging.info(f"Lunghezza massima strip: {lunghezza_massima_strip}mm")

                if lunghezza_totale > 0:
                    quantita_strip_led = math.ceil(lunghezza_totale / (lunghezza_massima_strip * 1000))
        except Exception as e:
            logging.error(f"Errore nel recupero lunghezza strip: {str(e)}")

    modalita = configurazione.get('modalitaConfigurazione', '')
    
    if modalita == 'solo_strip':
        strip_led = configurazione.get('stripLedSelezionata', '')
        nome_commerciale = configurazione.get('nomeCommercialeStripLed', '')
        
        if nome_commerciale:
            codice_prodotto = nome_commerciale
        elif strip_led:
            codice_prodotto = strip_led
        else:
            codice_prodotto = 'Strip LED'
    else:
        profilo = configurazione.get('profiloSelezionato', '')
        codice_prodotto = profilo if profilo else 'Configurazione'

    prezzi = db.get_prezzi_configurazione(
        tuttiCodici['profilo'],
        tuttiCodici['stripLed'],
        tuttiCodici['alimentatore'],
        tuttiCodici['dimmer'],
        finitura_profilo=configurazione.get('finituraSelezionata'),
        lunghezza_profilo=configurazione.get('lunghezzaRichiesta'),
        temperatura_strip=configurazione.get('temperaturaSelezionata') or configurazione.get('temperaturaColoreSelezionata'),
        potenza_strip=configurazione.get('potenzaSelezionata'),
        quantita_profilo=configurazione.get('quantitaProfilo', 1),
        quantita_strip=configurazione.get('quantitaStripLed', 1),
        lunghezze_multiple=configurazione.get('lunghezzeMultiple'),
        tappi_selezionati=configurazione.get('tappiSelezionati'),
        quantita_tappi=configurazione.get('quantitaTappi', 0),
        diffusore_selezionato=configurazione.get('diffusoreSelezionato'),
        quantita_diffusore=configurazione.get('quantitaDiffusore', 0)
    )

    return jsonify({
        'success': True,
        'riepilogo': configurazione,
        'potenzaTotale': round(potenza_totale, 2),
        'codiceProdotto': codice_prodotto,
        'quantitaProfilo': quantita_profilo,
        'quantitaStripLed': quantita_strip_led,
        'lunghezzaMassimaProfilo': lunghezza_massima_profilo,
        'lunghezzaMassimaStripLed': lunghezza_massima_strip,
        'lunghezzaTotale': lunghezza_totale,
        'combinazioneProfiloOttimale': combinazione_ottimale,
        'prezzi': prezzi
    })

@app.route('/salva_configurazione', methods=['POST'])
def salva_configurazione():
    try:
        data = request.json
        
        saved = db.salva_configurazione(
            configurazione=data.get('configurazione'),
            codice_prodotto=data.get('codice_prodotto'),
            email=data.get('email'),
            telefono=data.get('telefono'),
            note=data.get('note')
        )
        
        if saved:
            try:
                excel_path = genera_excel_configurazione(
                    data.get('configurazione'),
                    data.get('codice_prodotto')
                )
                
                return jsonify({
                    'success': True,
                    'id': saved['id'],
                    'message': 'Configurazione salvata con successo',
                    'excel_path': excel_path
                })
            except Exception as excel_error:
                return jsonify({
                    'success': True,
                    'id': saved['id'],
                    'message': 'Configurazione salvata (errore generazione Excel)',
                    'error_detail': str(excel_error)
                })
        
        return jsonify({'success': False, 'message': 'Errore nel salvataggio'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/download_excel/<filename>')
def download_excel(filename):
    try:
        filepath = os.path.join(tempfile.gettempdir(), filename)
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 404

def genera_excel_configurazione(configurazione, codice_prodotto):
    wb = Workbook()
    ws = wb.active
    ws.title = "Configurazione REDO"
    
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="E83F34", end_color="E83F34", fill_type="solid")
    
    subheader_font = Font(bold=True, size=12)
    subheader_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Configurazione REDO - {codice_prodotto}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = "Data:"
    ws['B2'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    row = 4
    
    sezioni = [
        ("Profilo", ["categoria", "profilo", "tipologia", "finitura"]),
        ("Strip LED", ["strip", "temperatura", "potenza", "tensione", "ip"]),
        ("Alimentazione", ["tipoAlimentazione", "tipoAlimentatore", "potenzaAlimentatore"]),
        ("Dimmerazione", ["dimmer"]),
        ("Cablaggio", ["uscitaCavo", "lunghezzaCavoIngresso", "lunghezzaCavoUscita"]),
        ("Dimensioni", ["forma", "lunghezzaTotale", "lunghezze"])
    ]
    
    for sezione, campi in sezioni:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = sezione
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        row += 1
        
        for campo in campi:
            if campo in configurazione:
                valore = configurazione[campo]
                
                if isinstance(valore, dict):
                    valore = json.dumps(valore, ensure_ascii=False)
                elif isinstance(valore, list):
                    valore = ", ".join(map(str, valore))
                elif valore is None:
                    valore = "N/A"
                
                ws[f'A{row}'] = campo.replace('_', ' ').title()
                ws[f'B{row}'] = str(valore)
                ws.merge_cells(f'B{row}:D{row}')
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].border = border
                
                row += 1
        
        row += 1
    
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Calcoli e Note"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    row += 1
    
    if 'calcoliLunghezza' in configurazione:
        calcoli = configurazione['calcoliLunghezza']
        ws[f'A{row}'] = "Spazio Buio:"
        ws[f'B{row}'] = f"{calcoli.get('spazioBuio', 0)} mm"
        row += 1
        
        ws[f'A{row}'] = "Taglio Strip:"
        ws[f'B{row}'] = f"ogni {calcoli.get('taglioMinimo', 0)} mm"
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    filename = f"REDO_Config_{codice_prodotto}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)
    
    return filename

@app.route('/get_tipologie_strip_disponibili')
def get_tipologie_strip_disponibili():
    """Ottiene tutte le tipologie di strip disponibili nel database"""
    try:
        tipologie_data = db.supabase.table('strip_led').select('tipo').execute().data
        tipologie_uniche = list(set([t['tipo'] for t in tipologie_data if t['tipo']]))
        
        return jsonify({'success': True, 'tipologie': tipologie_uniche})
    except Exception as e:
        logging.error(f"Errore in get_tipologie_strip_disponibili: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_special_strip_disponibili')
def get_special_strip_disponibili():
    """Ottiene tutte le special strip disponibili nel database"""
    try:
        strips_data = db.supabase.table('strip_led').select('nome_commerciale, id').execute().data
        special_strips = set()

        special_keywords = {
            'XFLEX': ['XFLEX', 'FLEX'],
            'XSNAKE': ['XSNAKE', 'SNAKE', 'SNK'],
            'XMAGIS': ['XMAGIS', 'MAGIS', 'MG13X12', 'MG12X17'],
            'ZIG_ZAG': ['ZIGZAG', 'ZIG_ZAG', 'ZIG-ZAG']
        }

        for strip in strips_data:
            nome_commerciale = (strip.get('nome_commerciale') or '').upper()
            strip_id = (strip.get('id') or '').upper()
            
            for special_type, keywords in special_keywords.items():
                if any(keyword in nome_commerciale or keyword in strip_id for keyword in keywords):
                    special_strips.add(special_type)
                    break
        
        return jsonify({'success': True, 'special_strips': list(special_strips)})
    except Exception as e:
        logging.error(f"Errore in get_special_strip_disponibili: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_opzioni_strip_standalone', methods=['POST'])
def get_opzioni_strip_standalone():
    """Ottiene le tensioni disponibili dal database per tipologia e special strip"""
    try:
        data = request.json
        tipologia = data.get('tipologia')
        special = data.get('special')
        query = db.supabase.table('strip_led').select('tensione')
        
        if tipologia and tipologia != 'None':
            if tipologia == 'SPECIAL':
                if special:
                    if special == 'XMAGIS':
                        query = query.or_('nome_commerciale.ilike.%XMAGIS%,id.ilike.%XMAGIS%,nome_commerciale.ilike.%MG13X12%,nome_commerciale.ilike.%MG12X17%')
                    elif special == 'XFLEX':
                        query = query.or_('nome_commerciale.ilike.%XFLEX%,id.ilike.%XFLEX%')
                    elif special == 'XSNAKE':
                        query = query.or_('nome_commerciale.ilike.%XSNAKE%,id.ilike.%XSNAKE%,id.ilike.%SNK%')
                    elif special == 'ZIG_ZAG':
                        query = query.or_('nome_commerciale.ilike.%ZIGZAG%,id.ilike.%ZIGZAG%,nome_commerciale.ilike.%ZIG_ZAG%')
                else:
                    special_keywords = ['XFLEX', 'XSNAKE', 'XMAGIS', 'ZIGZAG', 'ZIG_ZAG', 'MG13X12', 'MG12X17', 'SNK']
                    or_conditions = []
                    for keyword in special_keywords:
                        or_conditions.extend([f"nome_commerciale.ilike.%{keyword}%", f"id.ilike.%{keyword}%"])
                    query = query.or_(','.join(or_conditions))
            else:
                query = query.eq('tipo', tipologia)
        
        strips = query.execute().data
        
        if not strips:
            return jsonify({'success': True, 'tensioni': []})

        tensioni_uniche = list(set([s['tensione'] for s in strips if s['tensione']]))

        def get_tensione_order(tensione):
            try:
                return int(tensione.replace('V', ''))
            except:
                return 999
        
        tensioni_ordinate = sorted(tensioni_uniche, key=get_tensione_order)
        
        return jsonify({'success': True, 'tensioni': tensioni_ordinate})
    except Exception as e:
        logging.error(f"Errore in get_opzioni_strip_standalone: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_opzioni_ip_standalone', methods=['POST'])
def get_opzioni_ip_standalone():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'Nessun dato ricevuto'})
        
        tipologia = data.get('tipologia', '')
        tensione = data.get('tensione', '')
        special = data.get('special')

        query = db.supabase.table('strip_led').select('ip')\
            .eq('tensione', tensione)

        if tipologia and tipologia != 'None':
            query = query.eq('tipo', tipologia)

        if tipologia == 'SPECIAL' and special:
            query = db.supabase.table('strip_led').select('ip')\
                .eq('tensione', tensione)
            
            if special == 'XMAGIS':
                query = query.or_('nome_commerciale.ilike.%XMAGIS%,id.ilike.%XMAGIS%,nome_commerciale.ilike.%MG13X12%,nome_commerciale.ilike.%MG12X17%')
            elif special == 'XFLEX':
                query = query.or_('nome_commerciale.ilike.%XFLEX%,id.ilike.%XFLEX%')
            elif special == 'XSNAKE':
                query = query.or_('nome_commerciale.ilike.%XSNAKE%,id.ilike.%XSNAKE%,id.ilike.%SNK%')
            elif special == 'ZIG_ZAG':
                query = query.or_('nome_commerciale.ilike.%ZIGZAG%,id.ilike.%ZIGZAG%,nome_commerciale.ilike.%ZIG_ZAG%')

        strips = query.execute().data
        
        if not strips:
            logging.warning("DEBUG IP - Nessuna strip trovata per i parametri specificati")
            return jsonify({'success': True, 'gradi_ip': []})

        ip_disponibili = list(set([strip['ip'] for strip in strips if strip['ip']]))

        def get_ip_order(ip):
            ip_order = {'IP20': 1, 'IP44': 2, 'IP65': 3, 'IP66': 4, 'IP67': 5}
            return ip_order.get(ip, 999)
        
        ip_ordinati = sorted(ip_disponibili, key=get_ip_order)
        
        return jsonify({
            'success': True, 
            'gradi_ip': ip_ordinati,
            'debug': {
                'strips_trovate': len(strips),
                'ip_unici': len(ip_disponibili),
                'parametri': {
                    'tipologia': tipologia,
                    'tensione': tensione,
                    'special': special
                }
            }
        })
        
    except Exception as e:
        logging.error(f"DEBUG IP - ERRORE: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Errore interno: {str(e)}'})

@app.route('/get_opzioni_temperatura_standalone', methods=['POST'])
def get_opzioni_temperatura_standalone():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'Nessun dato ricevuto'})
        
        tipologia = data.get('tipologia', '')
        tensione = data.get('tensione', '')
        ip = data.get('ip', '')
        special = data.get('special')
        categoria = data.get('categoria')

        strips = []
        
        if tipologia == 'SPECIAL' and special:
            query = db.supabase.table('strip_led')\
                .select('id, nome_commerciale, tensione, ip, tipo')\
                .eq('tensione', tensione)\
                .eq('ip', ip)
            
            all_strips = query.execute().data

            special_keywords = {
                'XMAGIS': ['XMAGIS', 'MAGIS', 'MG13X12', 'MG12X17'],
                'XFLEX': ['XFLEX', 'FLEX'],
                'XSNAKE': ['XSNAKE', 'SNAKE', 'SNK'],
                'ZIG_ZAG': ['ZIGZAG', 'ZIG_ZAG', 'ZIG-ZAG'],
            }
            
            keywords = special_keywords.get(special, [])
            
            for strip in all_strips:
                nome_commerciale = (strip.get('nome_commerciale') or '').upper()
                strip_id = (strip.get('id') or '').upper()

                if any(keyword in nome_commerciale or keyword in strip_id for keyword in keywords):
                    strips.append(strip)
            
        else:
            query = db.supabase.table('strip_led').select('id, nome_commerciale, tensione, ip, tipo')\
                .eq('tensione', tensione)\
                .eq('ip', ip)
            
            if tipologia and tipologia not in ['None', 'SPECIAL']:
                query = query.eq('tipo', tipologia)
            
            strips = query.execute().data
        
        if not strips:
            logging.warning("DEBUG TEMPERATURA - Nessuna strip trovata per i parametri specificati")
            return jsonify({'success': True, 'temperature': []})

        strip_ids = [s['id'] for s in strips]
        temperature_data = db.supabase.table('strip_temperature')\
            .select('temperatura')\
            .in_('strip_id', strip_ids)\
            .execute().data
        
        if not temperature_data:
            logging.warning("DEBUG TEMPERATURA - Nessuna temperatura trovata per le strip")
            return jsonify({'success': True, 'temperature': []})

        temperature_uniche = list(set([t['temperatura'] for t in temperature_data if t['temperatura']]))

        def get_temperatura_order(temp):
            if 'K' in temp and temp.replace('K', '').isdigit():
                return (0, int(temp.replace('K', '')))
            elif temp == 'CCT':
                return (1, 0)
            elif temp == 'RGB':
                return (2, 0)
            elif temp == 'RGBW':
                return (3, 0)
            else:
                return (4, 0)
        
        temperature_ordinate = sorted(temperature_uniche, key=get_temperatura_order)
        
        return jsonify({
            'success': True, 
            'temperature': temperature_ordinate,
            'debug': {
                'strips_trovate': len(strips),
                'temperature_records': len(temperature_data),
                'temperature_uniche': len(temperature_uniche),
                'parametri': {
                    'tipologia': tipologia,
                    'tensione': tensione,
                    'ip': ip,
                    'special': special,
                    'categoria': categoria
                }
            }
        })
        
    except Exception as e:
        logging.error(f"DEBUG TEMPERATURA - ERRORE: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Errore interno: {str(e)}'})

@app.route('/get_opzioni_potenza_standalone', methods=['POST'])
def get_opzioni_potenza_standalone():
    try:
        data = request.json
        tipologia = data.get('tipologia')
        special = data.get('special')
        tensione = data.get('tensione')
        ip = data.get('ip')
        temperatura = data.get('temperatura')

        query = db.supabase.table('strip_led').select('id')\
            .eq('tensione', tensione)\
            .eq('ip', ip)

        if tipologia == 'SPECIAL' and special:
            if special == 'XMAGIS':
                query = query.or_('nome_commerciale.ilike.%XMAGIS%,id.ilike.%XMAGIS%,nome_commerciale.ilike.%MG13X12%,nome_commerciale.ilike.%MG12X17%')
            elif special == 'XFLEX':
                query = query.or_('nome_commerciale.ilike.%XFLEX%,id.ilike.%XFLEX%')
            elif special == 'XSNAKE':
                query = query.or_('nome_commerciale.ilike.%XSNAKE%,id.ilike.%XSNAKE%,id.ilike.%SNK%')
            elif special == 'ZIG_ZAG':
                query = query.or_('nome_commerciale.ilike.%ZIGZAG%,id.ilike.%ZIGZAG%,nome_commerciale.ilike.%ZIG_ZAG%')
        elif tipologia and tipologia != 'SPECIAL':
            query = query.eq('tipo', tipologia)

        strips = query.execute().data
        
        if not strips:
            return jsonify({'success': True, 'potenze': []})
        
        strip_ids = [s['id'] for s in strips]

        if temperatura:
            temp_check = db.supabase.table('strip_temperature')\
                .select('strip_id')\
                .eq('temperatura', temperatura)\
                .in_('strip_id', strip_ids)\
                .execute().data
            
            strip_ids = [t['strip_id'] for t in temp_check]
        
        if not strip_ids:
            return jsonify({'success': True, 'potenze': []})

        potenze_data = db.supabase.table('strip_potenze')\
            .select('potenza')\
            .in_('strip_id', strip_ids)\
            .execute().data

        potenze_uniche = sorted(list(set([p['potenza'] for p in potenze_data])), 
                               key=lambda x: float(x.replace('W/m', '').replace(',', '.').split()[0]))
        
        potenze = [{'id': p, 'nome': p} for p in potenze_uniche]
        
        return jsonify({'success': True, 'potenze': potenze})
        
    except Exception as e:
        logging.error(f"Errore in get_opzioni_potenza_standalone: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_strip_led_filtrate_standalone', methods=['POST'])
def get_strip_led_filtrate_standalone():
    try:
        data = request.json

        tipologia = data.get('tipologia')
        special = data.get('special')
        tensione = data.get('tensione')
        ip = data.get('ip')
        temperatura = data.get('temperatura')
        potenza = data.get('potenza')

        if tipologia == 'SPECIAL' and special:
            
            special_keywords = {
                'XFLEX': ['XFLEX', 'FLEX'],
                'XSNAKE': ['XSNAKE', 'SNAKE'],
                'XMAGIS': ['XMAGIS', 'MAGIS'],
                'ZIG_ZAG': ['ZIGZAG', 'ZIG_ZAG'],
            }
            
            keywords = special_keywords.get(special, [])
            
            if not keywords:
                return jsonify({'success': False, 'message': f'Tipologia special strip non riconosciuta: {special}'})

            base_query = db.supabase.table('strip_led').select('*')
            
            or_conditions = []
            for keyword in keywords:
                or_conditions.append(f"nome_commerciale.ilike.%{keyword}%")
                or_conditions.append(f"id.ilike.%{keyword}%")
            
            or_query = ','.join(or_conditions)
            
            special_strips = base_query.or_(or_query).execute().data

            if tensione:
                special_strips = [s for s in special_strips if s['tensione'] == tensione]

            if ip:
                strips_con_ip_richiesto = [s for s in special_strips if s['ip'] == ip]
                if strips_con_ip_richiesto:
                    special_strips = strips_con_ip_richiesto

            strips = special_strips
            
        else:
            query = db.supabase.table('strip_led').select('*')
            
            if tensione:
                query = query.eq('tensione', tensione)
            if ip:
                query = query.eq('ip', ip)
                
            strips = query.execute().data

        
        result = []
        for strip in strips:
            strip_id = strip['id']

            if temperatura:
                temp_check = db.supabase.table('strip_temperature')\
                    .select('temperatura')\
                    .eq('strip_id', strip_id)\
                    .eq('temperatura', temperatura)\
                    .execute().data
                
                if not temp_check:
                    continue

            if potenza:
                potenza_check = db.supabase.table('strip_potenze')\
                    .select('potenza')\
                    .eq('strip_id', strip_id)\
                    .eq('potenza', potenza)\
                    .execute().data

                if not potenza_check:
                    continue

            temperatures = db.supabase.table('strip_temperature')\
                .select('temperatura')\
                .eq('strip_id', strip_id)\
                .execute().data

            potenze = db.supabase.table('strip_potenze')\
                .select('potenza, codice_prodotto')\
                .eq('strip_id', strip_id)\
                .order('indice')\
                .execute().data
            
            strip['temperaturaColoreDisponibili'] = [t['temperatura'] for t in temperatures]
            strip['potenzeDisponibili'] = [p['potenza'] for p in potenze]
            strip['codiciProdotto'] = [p['codice_prodotto'] for p in potenze]
            strip['nomeCommerciale'] = strip.get('nome_commerciale', '')
            strip['taglioMinimo'] = strip.get('taglio_minimo', {})
            strip['temperatura'] = temperatura

            result.append(strip)

        return jsonify({'success': True, 'strip_led': result})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_strip_compatibili_esterni/<categoria>')
def get_strip_compatibili_esterni(categoria):
    try:
        profili = db.get_profili_by_categoria(categoria)

        strip_compatibili_totali = set()
        for profilo in profili:
            if 'stripLedCompatibili' in profilo:
                strip_compatibili_totali.update(profilo['stripLedCompatibili'])

        if strip_compatibili_totali:
            strip_details = db.supabase.table('strip_led')\
                .select('*')\
                .in_('id', list(strip_compatibili_totali))\
                .execute()
            
            return jsonify({
                'success': True,
                'strip_compatibili': list(strip_compatibili_totali),
                'strip_details': strip_details.data if strip_details.data else []
            })
        else:
            return jsonify({
                'success': True,
                'strip_compatibili': [],
                'strip_details': []
            })
            
    except Exception as e:
        logging.error(f"Errore in get_strip_compatibili_esterni: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/get_strip_led_by_nome_commerciale/<nome_commerciale>')
def get_strip_led_by_nome_commerciale(nome_commerciale):
    try:
        strip_data = db.supabase.table('strip_led').select('*').eq('nome_commerciale', nome_commerciale).single().execute()
        
        if not strip_data.data:
            return jsonify({'success': False, 'message': f'Strip LED non trovata: {nome_commerciale}'})
        
        strip_info = strip_data.data
        
        temperature = db.supabase.table('strip_temperature').select('temperatura').eq('strip_id', strip_info['id']).execute().data
        potenze = db.supabase.table('strip_potenze').select('*').eq('strip_id', strip_info['id']).order('indice').execute().data
        
        return jsonify({
            'success': True,
            'strip_led': {
                'id': strip_info['id'],
                'nome': strip_info['nome'],
                'nomeCommerciale': strip_info.get('nome_commerciale', ''),
                'tensione': strip_info['tensione'],
                'ip': strip_info['ip'],
                'temperaturaColoreDisponibili': [t['temperatura'] for t in temperature],
                'potenzeDisponibili': [p['potenza'] for p in potenze],
                'codiciProdotto': [p['codice_prodotto'] for p in potenze]
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'database': 'supabase'})

@app.route('/richiedi_preventivo', methods=['POST'])
def richiedi_preventivo():
    try:
        data = request.json

        nome_agente = data.get('nomeAgente', '')
        email_agente = data.get('emailAgente', '')
        ragione_sociale = data.get('ragioneSociale', '')
        riferimento = data.get('riferimento', '')
        note = data.get('note', '')
        configurazione = data.get('configurazione', {})
        codice_prodotto = data.get('codiceProdotto', '')

        if not nome_agente or not email_agente or not riferimento:
            return jsonify({
                'success': False, 
                'message': 'Nome agente, email agente e riferimento sono obbligatori'
            }), 400

        email_html = genera_email_preventivo(
            nome_agente, email_agente, ragione_sociale, 
            riferimento, note, configurazione, codice_prodotto
        )

        success = invia_email_preventivo(
            email_html, nome_agente, email_agente, 
            ragione_sociale, riferimento, codice_prodotto
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Preventivo inviato con successo'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Errore nell\'invio dell\'email'
            }), 500
            
    except Exception as e:
        logging.error(f"Errore in richiedi_preventivo: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Errore interno del server'
        }), 500

def invia_email_preventivo(email_html, nome_agente, email_agente, ragione_sociale, riferimento, codice_prodotto):
    """Invia l'email del preventivo"""
    try:
        if app.config.get('MAIL_USERNAME') and app.config.get('MAIL_PASSWORD'):
            return invia_con_flask_mail(email_html, nome_agente, email_agente, ragione_sociale, riferimento, codice_prodotto)
        else:
            return salva_preventivo_log(email_html, nome_agente, email_agente, ragione_sociale, riferimento, codice_prodotto)
            
    except Exception as e:
        logging.error(f"Errore invio email: {str(e)}")
        return False

def invia_con_flask_mail(email_html, nome_agente, email_agente, ragione_sociale, riferimento, codice_prodotto):
    """Invia email usando Flask-Mail"""
    try:
        subject = f"Richiesta Preventivo REDO - {codice_prodotto} - {riferimento}"
        
        msg = Message(
            subject=subject,
            recipients=['furlaninicoletta@gmail.com', 'realizzazioni@redogroup.it', email_agente],
            html=email_html,
            sender=app.config['MAIL_DEFAULT_SENDER']
        )
        
        mail.send(msg)
        logging.info(f"Email preventivo inviata con successo per {nome_agente}")
        return True
        
    except Exception as e:
        logging.error(f"Errore Flask-Mail: {str(e)}")
        return False

def salva_preventivo_log(email_html, nome_agente, email_agente, ragione_sociale, riferimento, codice_prodotto):
    """Fallback: salva il preventivo in un file di log"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"preventivo_{timestamp}_{riferimento}.html"
        filepath = os.path.join(os.path.dirname(__file__), 'preventivi_log', filename)

        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(email_html)
        
        logging.info(f"Preventivo salvato in: {filepath}")
        return True
        
    except Exception as e:
        logging.error(f"Errore salvataggio log: {str(e)}")
        return False

def genera_email_preventivo(nome_agente, email_agente, ragione_sociale, riferimento, note, configurazione, codice_prodotto):
    """Genera l'HTML dell'email del preventivo con tutti i dati del riepilogo"""

    mappaCategorieVisualizzazione = {
        'nanoprofili': 'Nanoprofili',
        'incasso': 'Profili a Incasso',
        'sospensione': 'Profili a Sospensione',
        'plafone': 'Profili a Plafone',
        'parete': 'Profili a Parete',
        'particolari': 'Profili Particolari',
        'scalino': 'Profili a Scalino',
        'wall_washer': 'Profili Wallwasher',
        'wall_washer_ext': 'Profili Wallwasher da Esterni',
        'esterni': 'Profili per Strip LED da Esterni',
    }
    
    mappaTipologieVisualizzazione = {
        'taglio_misura': 'Taglio su misura',
        'profilo_intero': 'Profilo intero'
    }
    
    mappaStripLedVisualizzazione = {
        'senza_strip': 'Senza Strip LED',
        'STRIP_24V_SMD_IP20': 'STRIP 24V SMD (IP20)',
        'STRIP_24V_COB_IP20_HIGH': 'STRIP 24V COB (IP20) HIGH POWER',
        'STRIP_24V_SMD_IP66': 'STRIP 24V SMD (IP66)',
        'STRIP_24V_COB_IP20': 'STRIP 24V COB (IP20)',
        'STRIP_24V_COB_IP66': 'STRIP 24V COB (IP66)',
        'STRIP_48V_SMD_IP20': 'STRIP 48V SMD (IP20)',
        'STRIP_48V_SMD_IP66': 'STRIP 48V SMD (IP66)',
        'STRIP_24V_RGB_SMD_IP20': 'STRIP 24V RGB SMD (IP20)',
        'STRIP_24V_RGB_SMD_IP66': 'STRIP 24V RGB SMD (IP66)',
        'STRIP_24V_RGB_COB_IP20': 'STRIP 24V RGB COB (IP20)',
        'STRIP_24V_RGB_COB_IP66': 'STRIP 24V RGB COB (IP66)',
        'STRIP_220V_COB_IP20': 'STRIP 220V COB (IP20)',
        'STRIP_220V_COB_IP66': 'STRIP 220V COB (IP66)',
    }
    
    mappaFormeTaglio = {
        'DRITTO_SEMPLICE': 'Dritto semplice',
        'FORMA_L_DX': 'Forma a L DX',
        'FORMA_L_SX': 'Forma a L SX',
        'FORMA_C': 'Forma a C',
        'RETTANGOLO_QUADRATO': 'Rettangolo/Quadrato'
    }
    
    mappaFiniture = {
        'ALLUMINIO_ANODIZZATO': 'Alluminio anodizzato',
        'BIANCO': 'Bianco',
        'NERO': 'Nero',
        'ALLUMINIO': 'Alluminio'
    }
    
    mappaTipologiaStripVisualizzazione = {
        'COB': 'COB (Chip On Board)',
        'SMD': 'SMD (Surface Mount Device)',
        'SPECIAL': 'Special Strip'
    }
    
    mappaSpecialStripVisualizzazione = {
        'XFLEX': 'XFLEX',
        'ZIG_ZAG': 'ZIG ZAG',
        'XSNAKE': 'XSNAKE',
        'XMAGIS': 'XMAGIS'
    }
    
    data_corrente = datetime.now().strftime("%d/%m/%Y %H:%M")

    def calcola_codici_prodotto():
        codici = {
            'profilo': '',
            'stripLed': '',
            'alimentatore': '',
            'dimmer': ''
        }
        
        if configurazione.get('codiceProfilo'):
            codici['profilo'] = configurazione['codiceProfilo']
        elif configurazione.get('profiloSelezionato'):
            codici['profilo'] = configurazione['profiloSelezionato'].replace('_', '/')
            
        if configurazione.get('stripLedSelezionata'):
            codici['stripLed'] = configurazione['stripLedSelezionata']
        elif configurazione.get('stripLedSceltaFinale'):
            codici['stripLed'] = configurazione['stripLedSceltaFinale']
            
        if configurazione.get('codiceAlimentatore'):
            codici['alimentatore'] = configurazione['codiceAlimentatore']
            
        if configurazione.get('codiceDimmer'):
            codici['dimmer'] = configurazione['codiceDimmer']
        elif configurazione.get('dimmerSelezionato') == 'NESSUN_DIMMER':
            codici['dimmer'] = ''
        elif configurazione.get('dimmerCodice'):
            codici['dimmer'] = configurazione['dimmerCodice']
            
        return codici
    
    tuttiCodici = calcola_codici_prodotto()

    prezzi = db.get_prezzi_configurazione(
        tuttiCodici['profilo'],
        tuttiCodici['stripLed'],
        tuttiCodici['alimentatore'],
        tuttiCodici['dimmer'],
        finitura_profilo=configurazione.get('finituraSelezionata'),
        lunghezza_profilo=configurazione.get('lunghezzaRichiesta'),
        temperatura_strip=configurazione.get('temperaturaSelezionata') or configurazione.get('temperaturaColoreSelezionata'),
        potenza_strip=configurazione.get('potenzaSelezionata'),
        quantita_profilo=configurazione.get('quantitaProfilo', 1),
        quantita_strip=configurazione.get('quantitaStripLed', 1),
        lunghezze_multiple=configurazione.get('lunghezzeMultiple'),
        tappi_selezionati=configurazione.get('tappiSelezionati'),
        quantita_tappi=configurazione.get('quantitaTappi', 0),
        diffusore_selezionato=configurazione.get('diffusoreSelezionato'),
        quantita_diffusore=configurazione.get('quantitaDiffusore', 0)
    )

    lunghezza_cavo_totale_mm = 0
    if configurazione.get('lunghezzaCavoIngresso'):
        lunghezza_cavo_totale_mm += configurazione['lunghezzaCavoIngresso']
    if configurazione.get('lunghezzaCavoUscita'):
        lunghezza_cavo_totale_mm += configurazione['lunghezzaCavoUscita']

    lunghezza_cavo_totale_metri_esatti = lunghezza_cavo_totale_mm / 1000 if lunghezza_cavo_totale_mm > 0 else 0
    metri_da_fatturare = math.ceil(lunghezza_cavo_totale_metri_esatti) if lunghezza_cavo_totale_mm > 0 else 0
    prezzo_cavo = metri_da_fatturare * 2

    def get_nome_visualizzabile(valore, mappa):
        return mappa.get(valore, valore) if valore else 'N/A'

    def format_prezzo(prezzo):
        if not prezzo or prezzo == 0:
            return ''
        return f' - €{prezzo:.2f}'
    
    def get_codici_dal_database():
        codici_email = {
            'profilo': '',
            'stripLed': ''
        }

        if configurazione.get('profiloSelezionato'):
            profilo_id = configurazione['profiloSelezionato']
            finitura = configurazione.get('finituraSelezionata')

            lunghezza = None
            if configurazione.get('combinazioneProfiloOttimale') and len(configurazione['combinazioneProfiloOttimale']) > 0:
                lunghezze_combinazione = [combo['lunghezza'] for combo in configurazione['combinazioneProfiloOttimale']]
                lunghezza = max(lunghezze_combinazione)
            else:
                lunghezza = configurazione.get('lunghezzaRichiesta')
            
            try:
                codici_email['profilo'] = db.get_codice_profilo(profilo_id, finitura, int(lunghezza) if lunghezza else None)
            except Exception as e:
                logging.error(f"Errore recupero codice profilo: {str(e)}")
                codici_email['profilo'] = profilo_id.replace('_', '/')

        strip_id = (configurazione.get('stripLedSelezionata') or 
            configurazione.get('stripLedSceltaFinale') or 
            configurazione.get('stripLedSelezionata'))

        if strip_id and strip_id not in ['NO_STRIP', 'senza_strip', '', None]:
            temperatura = (configurazione.get('temperaturaColoreSelezionata') or 
                        configurazione.get('temperaturaSelezionata'))
            potenza = configurazione.get('potenzaSelezionata')
            
            try:
                codici_email['stripLed'] = db.get_codice_strip_led(strip_id, temperatura, potenza)
            except Exception as e:
                logging.error(f"Errore recupero codice strip: {str(e)}")
                codici_email['stripLed'] = strip_id
        
        return codici_email

    codici_email = get_codici_dal_database()
    
    html = f"""
    <!DOCTYPE html>
    <html lang="it">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Richiesta Preventivo REDO</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 800px;
                margin: 0 auto;
                padding: 20px;
            }}
            .header {{
                background-color: #e83f34;
                color: white;
                padding: 20px;
                text-align: center;
                margin-bottom: 30px;
            }}
            .section {{
                margin-bottom: 25px;
                padding: 15px;
                border: 1px solid #ddd;
                border-radius: 5px;
            }}
            .section h3 {{
                color: #e83f34;
                border-bottom: 2px solid #e83f34;
                padding-bottom: 5px;
                margin-top: 0;
            }}
            .data-table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .data-table th, .data-table td {{
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }}
            .data-table th {{
                background-color: #f5f5f5;
                font-weight: bold;
            }}
            .alert-warning {{
                background-color: #fff3cd;
                border: 1px solid #ffeaa7;
                color: #856404;
                padding: 10px;
                border-radius: 5px;
                margin-top: 15px;
            }}
            .footer {{
                text-align: center;
                margin-top: 30px;
                padding-top: 20px;
                border-top: 1px solid #ddd;
                color: #666;
                font-size: 0.9em;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Richiesta Preventivo REDO</h1>
            <h2>Codice Prodotto: {codice_prodotto}</h2>
            <p>Data richiesta: {data_corrente}</p>
        </div>

        <div class="section">
            <h3>Dati Agente</h3>
            <table class="data-table">
                <tr><th>Nome Agente</th><td>{nome_agente}</td></tr>
                <tr><th>Email Agente</th><td>{email_agente}</td></tr>
                <tr><th>Ragione Sociale</th><td>{ragione_sociale or 'Non specificata'}</td></tr>
                <tr><th>Riferimento</th><td>{riferimento}</td></tr>
                {f'<tr><th>Note</th><td>{note}</td></tr>' if note else ''}
            </table>
        </div>

        <div class="section">
            <h3>Configurazione Prodotto</h3>
            <table class="data-table">
    """

    if configurazione.get('categoriaSelezionata'):
        html += f"<tr><th>Categoria</th><td>{get_nome_visualizzabile(configurazione['categoriaSelezionata'], mappaCategorieVisualizzazione)}</td></tr>"

    if configurazione.get('nomeModello'):
        modello_text = configurazione['nomeModello']

        if codici_email['profilo']:
            modello_text += f" - {codici_email['profilo']}"

        if configurazione.get('quantitaProfilo', 1) > 1:
            if configurazione.get('combinazioneProfiloOttimale'):
                parti = []
                for combo in configurazione['combinazioneProfiloOttimale']:
                    if combo['quantita'] > 1:
                        parte = f"{combo['quantita']}x {modello_text} ({combo['lunghezza']}mm cad.)"
                    else:
                        parte = f"{modello_text} ({combo['lunghezza']}mm)"
                    parti.append(parte)
                modello_text = " + ".join(parti)
            else:
                modello_text = f"{configurazione['quantitaProfilo']}x {modello_text}"

        prezzo_profilo_text = format_prezzo(prezzi['profilo'])
        modello_text += prezzo_profilo_text
        
        html += f"<tr><th>Modello</th><td>{modello_text}</td></tr>"

    if configurazione.get('tipologiaSelezionata'):
        html += f"<tr><th>Tipologia</th><td>{get_nome_visualizzabile(configurazione['tipologiaSelezionata'], mappaTipologieVisualizzazione)}</td></tr>"

    if configurazione.get('lunghezzaRichiesta'):
        html += f"<tr><th>Lunghezza richiesta</th><td>{configurazione['lunghezzaRichiesta']}mm</td></tr>"

    if configurazione.get('lunghezzeMultiple') and configurazione.get('formaDiTaglioSelezionata') != 'DRITTO_SEMPLICE':
        etichette_lati = {
            'FORMA_L_DX': {'lato1': 'Lato orizzontale', 'lato2': 'Lato verticale'},
            'FORMA_L_SX': {'lato1': 'Lato orizzontale', 'lato2': 'Lato verticale'},
            'FORMA_C': {'lato1': 'Lato orizzontale superiore', 'lato2': 'Lato verticale', 'lato3': 'Lato orizzontale inferiore'},
            'RETTANGOLO_QUADRATO': {'lato1': 'Lunghezza', 'lato2': 'Larghezza'}
        }
        
        etichette = etichette_lati.get(configurazione.get('formaDiTaglioSelezionata'), {})
        
        for lato, valore in configurazione['lunghezzeMultiple'].items():
            if valore:
                etichetta = etichette.get(lato, f"Lato {lato.replace('lato', '')}")
                html += f"<tr><th>{etichetta}</th><td>{valore}mm</td></tr>"

    if configurazione.get('stripLedSelezionata') and configurazione['stripLedSelezionata'] not in ['NO_STRIP', 'senza_strip']:
        nome_strip = configurazione.get('nomeCommercialeStripLed') or get_nome_visualizzabile(configurazione['stripLedSelezionata'], mappaStripLedVisualizzazione)

        if configurazione.get('quantitaStripLed', 1) > 1:
            nome_strip = f"{configurazione['quantitaStripLed']}x {nome_strip} ({configurazione.get('lunghezzaMassimaStripLed', 5) * 1000}mm cad.)"

        if codici_email['stripLed']:
            nome_strip += f" - {codici_email['stripLed']}"

        prezzo_strip_text = format_prezzo(prezzi['strip_led'])
        nome_strip += prezzo_strip_text

        html += f"<tr><th>Strip LED</th><td>{nome_strip}</td></tr>"
    else:
        html += f"<tr><th>Strip LED</th><td>Senza Strip LED</td></tr>"

    if configurazione.get('tipologiaStripSelezionata'):
        tipologia_strip_text = get_nome_visualizzabile(configurazione['tipologiaStripSelezionata'], mappaTipologiaStripVisualizzazione)
        if configurazione['tipologiaStripSelezionata'] == 'SPECIAL' and configurazione.get('specialStripSelezionata'):
            tipologia_strip_text += f" - {get_nome_visualizzabile(configurazione['specialStripSelezionata'], mappaSpecialStripVisualizzazione)}"
        html += f"<tr><th>Tipologia Strip</th><td>{tipologia_strip_text}</td></tr>"

    if configurazione.get('potenzaSelezionata'):
        html += f"<tr><th>Potenza</th><td>{configurazione['potenzaSelezionata']}</td></tr>"

    if configurazione.get('tensioneSelezionato') == '220V':
        html += f"<tr><th>Alimentazione</th><td>Strip 220V (no alimentatore)</td></tr>"
    elif configurazione.get('alimentazioneSelezionata'):
        alimentazione_text = 'ON/OFF' if configurazione['alimentazioneSelezionata'] == 'ON-OFF' else configurazione['alimentazioneSelezionata'].replace('_', ' ')
        html += f"<tr><th>Alimentazione</th><td>{alimentazione_text}</td></tr>"

        if configurazione.get('tipologiaAlimentatoreSelezionata') and configurazione['alimentazioneSelezionata'] != 'SENZA_ALIMENTATORE':
            alimentatore_text = configurazione['tipologiaAlimentatoreSelezionata']
            if tuttiCodici['alimentatore']:
                alimentatore_text += f" - {tuttiCodici['alimentatore']}"
            prezzo_alimentatore_text = format_prezzo(prezzi['alimentatore'])
            alimentatore_text += prezzo_alimentatore_text
            html += f"<tr><th>Alimentatore</th><td>{alimentatore_text}</td></tr>"

        if configurazione.get('potenzaConsigliataAlimentatore') and configurazione.get('tensioneSelezionato') != '220V':
            html += f"<tr><th>Potenza consigliata</th><td>{configurazione['potenzaConsigliataAlimentatore']}W</td></tr>"

    if configurazione.get('dimmerSelezionato'):
        if configurazione.get('tensioneSelezionato') == '220V' and configurazione['dimmerSelezionato'] == 'DIMMER_A_PULSANTE_SEMPLICE':
            dimmer_text = 'CTR130 - Dimmerabile TRIAC tramite pulsante e sistema TUYA'
        else:
            dimmer_text = 'Nessun dimmer' if configurazione['dimmerSelezionato'] == 'NESSUN_DIMMER' else configurazione['dimmerSelezionato'].replace('_', ' ')
            if tuttiCodici['dimmer'] and configurazione['dimmerSelezionato'] != 'NESSUN_DIMMER':
                dimmer_text += f" {tuttiCodici['dimmer']}"
        prezzo_dimmer_text = format_prezzo(prezzi['dimmer'])
        dimmer_text += prezzo_dimmer_text
        
        html += f"<tr><th>Dimmer</th><td>{dimmer_text}</td></tr>"

        if configurazione.get('dimmerSelezionato') and 'TOUCH_SU_PROFILO' in configurazione['dimmerSelezionato']:
            html += f"<tr><th>Nota dimmer</th><td class='text-warning'>Spazio non illuminato di 50mm per touch su profilo</td></tr>"

    if configurazione.get('tipoAlimentazioneCavo'):
        alimentazione_cavo_text = 'Alimentazione unica' if configurazione['tipoAlimentazioneCavo'] == 'ALIMENTAZIONE_UNICA' else 'Alimentazione doppia'
        html += f"<tr><th>Alimentazione cavo</th><td>{alimentazione_cavo_text}</td></tr>"

        if configurazione.get('lunghezzaCavoIngresso'):
            html += f"<tr><th>Lunghezza cavo ingresso</th><td>{configurazione['lunghezzaCavoIngresso']}mm</td></tr>"

        if configurazione['tipoAlimentazioneCavo'] == 'ALIMENTAZIONE_DOPPIA' and configurazione.get('lunghezzaCavoUscita'):
            html += f"<tr><th>Lunghezza cavo uscita</th><td>{configurazione['lunghezzaCavoUscita']}mm</td></tr>"

    # 🔧 AGGIUNTA: Lunghezza cavo totale
    if lunghezza_cavo_totale_mm > 0:
        html += f"<tr><th>Lunghezza cavo</th><td>{lunghezza_cavo_totale_mm}mm</td></tr>"

    if configurazione.get('uscitaCavoSelezionata') and configurazione.get('categoriaSelezionata') not in ['esterni', 'wall_washer_ext']:
        uscita_cavo_text = configurazione['uscitaCavoSelezionata']
        if uscita_cavo_text == 'DRITTA':
            uscita_cavo_text = 'Dritta'
        elif uscita_cavo_text == 'LATERALE_DX':
            uscita_cavo_text = 'Laterale destra'
        elif uscita_cavo_text == 'LATERALE_SX':
            uscita_cavo_text = 'Laterale sinistra'
        elif uscita_cavo_text == 'RETRO':
            uscita_cavo_text = 'Retro'
        
        html += f"<tr><th>Uscita cavo</th><td>{uscita_cavo_text}</td></tr>"

    if configurazione.get('formaDiTaglioSelezionata'):
        html += f"<tr><th>Forma di taglio</th><td>{get_nome_visualizzabile(configurazione['formaDiTaglioSelezionata'], mappaFormeTaglio)}</td></tr>"

    if configurazione.get('finituraSelezionata'):
        html += f"<tr><th>Finitura</th><td>{get_nome_visualizzabile(configurazione['finituraSelezionata'], mappaFiniture)}</td></tr>"

    if configurazione.get('potenzaTotale'):
        html += f"<tr><th>Potenza totale</th><td>{configurazione['potenzaTotale']}W</td></tr>"

    # Only show tappi in riepilogo if they are not automatically included
    if configurazione.get('tappiSelezionati') and configurazione.get('quantitaTappi', 0) > 0 and not configurazione.get('tappiInclusi', False):
        tappo = configurazione['tappiSelezionati']
        quantita_selezionata = configurazione['quantitaTappi']
        prezzo_tappi = prezzi.get('tappi', 0)
        html += f"<tr><th>Tappi</th><td>{quantita_selezionata}x {tappo['codice']} - €{prezzo_tappi:.2f}</td></tr>"

    if configurazione.get('diffusoreSelezionato') and configurazione.get('quantitaDiffusore', 0) > 0:
        diffusore = configurazione['diffusoreSelezionato']
        quantita_diffusore = configurazione['quantitaDiffusore']
        prezzo_diffusore = prezzi.get('diffusore', 0)
        html += f"<tr><th>Diffusore</th><td>{quantita_diffusore}x {diffusore['codice']} - €{prezzo_diffusore:.2f}</td></tr>"

    # Show staffe with quantity and price in email
    if configurazione.get('staffaSelezionata'):
        staffa = configurazione['staffaSelezionata']
        quantita_staffe = configurazione.get('quantitaStaffe', configurazione.get('quantitaProfilo', 1))
        prezzo_unitario_staffa = float(staffa.get('prezzo', 0))
        prezzo_totale_staffe = prezzo_unitario_staffa * quantita_staffe
        html += f"<tr><th>Staffe di fissaggio</th><td>{quantita_staffe}x {staffa['codice']} - €{prezzo_totale_staffe:.2f}</td></tr>"

    if prezzi.get('totale', 0) > 0:
        html += f"""
                <tr style="border-top: 2px solid #e83f34; font-weight: bold;">
                    <th scope="row"><strong>Totale configurazione</strong></th>
                    <td><strong>€{prezzi['totale']:.2f}</strong></td>
                </tr>
        """

    html += """
            </table>
        </div>
    """

    html += """
        <div class="alert-warning">
            <strong>Note importanti:</strong><br>
            • Eventuali staffe aggiuntive non incluse<br>
    """

    if configurazione.get('categoriaSelezionata') in ['esterni', 'wall_washer_ext']:
        html += "• La lunghezza richiesta fa riferimento alla strip led esclusa di tappi e il profilo risulterà leggermente più corto<br>"
    else:
        html += "• Verrà aggiunto automaticamente uno spazio di 5mm per i tappi e la saldatura<br>"

    if configurazione.get('formaDiTaglioSelezionata') and configurazione['formaDiTaglioSelezionata'] != 'DRITTO_SEMPLICE':
        html += "• I profili verranno consegnati non assemblati tra di loro e la strip verrà consegnata non installata<br>"

    html += """
        </div>
    """

    costo_lavorazione_profilo = 0
    costo_taglio_strip = 0
    costo_gestione = 7

    if prezzi.get('totale', 0) > 0:

        if configurazione.get('tipologiaSelezionata') == 'taglio_misura':
            costo_lavorazione_profilo = 5

        if (configurazione.get('stripLedSelezionata') and 
            configurazione.get('stripLedSelezionata') not in ['NO_STRIP', 'senza_strip'] and
            configurazione.get('includeStripLed') != False):
            costo_taglio_strip = 5

        totale_base = prezzi['totale']
        totale_lavorazioni = costo_lavorazione_profilo + costo_taglio_strip + costo_gestione + prezzo_cavo
        totale_finale = totale_base + totale_lavorazioni
        
        html += f"""
        <div class="section">
            <h3>Riepilogo Prezzi</h3>
            <table class="data-table">
                <tr>
                    <th>TOTALE CONFIGURAZIONE</th>
                    <td>€{totale_base:.2f}</td>
                </tr>
        """

        if costo_lavorazione_profilo > 0:
            html += f"""
                <tr>
                    <th>Lavorazione profilo</th>
                    <td>€{costo_lavorazione_profilo:.2f}</td>
                </tr>
            """
        
        if costo_taglio_strip > 0:
            html += f"""
                <tr>
                    <th>Taglio e cablaggio strip</th>
                    <td>€{costo_taglio_strip:.2f}</td>
                </tr>
            """

        if prezzo_cavo > 0:
            html += f"""
                <tr>
                    <th>Prezzo cavo</th>
                    <td>€{prezzo_cavo:.2f}</td>
                </tr>
            """

    html += f"""
            <tr>
                <th>Spesa di gestione fissa</th>
                <td>€{costo_gestione:.2f}</td>
            </tr>
            <tr class="totale-row" style="border-top: 2px solid #e83f34;">
                <th><strong>TOTALE CONFIGURAZIONE E LAVORAZIONE</strong></th>
                <td><strong>€{totale_finale:.2f}</strong></td>
            </tr>
        </table>
    </div>
    """

    html += """
        <div class="footer">
            <p><strong>REDO Srl</strong> - Configuratore Profili LED</p>
            <p>Questa è una richiesta di preventivo generata automaticamente dal configuratore online.</p>
        </div>
    </body>
    </html>
    """
    
    return html

@app.route('/get_prezzi_configurazione', methods=['POST'])
def get_prezzi_configurazione():
    """Endpoint per ottenere i prezzi di una configurazione"""
    try:
        data = request.json
        
        codice_profilo = data.get('codice_profilo', '')
        codice_strip = data.get('codice_strip', '')
        codice_alimentatore = data.get('codice_alimentatore', '')
        codice_dimmer = data.get('codice_dimmer', '')

        finitura_profilo = data.get('finitura_profilo')
        lunghezza_profilo = data.get('lunghezza_profilo')
        temperatura_strip = data.get('temperatura_strip')
        potenza_strip = data.get('potenza_strip')

        quantita_profilo = data.get('quantita_profilo', 1)
        quantita_strip = data.get('quantita_strip', 1)

        prezzi = db.get_prezzi_configurazione(
            codice_profilo, 
            codice_strip, 
            codice_alimentatore, 
            codice_dimmer,
            finitura_profilo=finitura_profilo,
            lunghezza_profilo=lunghezza_profilo,
            temperatura_strip=temperatura_strip,
            potenza_strip=potenza_strip,
            quantita_profilo=quantita_profilo,
            quantita_strip=quantita_strip
        )
        
        return jsonify({
            'success': True,
            'prezzi': prezzi
        })
        
    except Exception as e:
        logging.error(f"Errore in get_prezzi_configurazione: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e),
            'prezzi': {
                'profilo': 0.0,
                'strip_led': 0.0,
                'alimentatore': 0.0,
                'dimmer': 0.0,
                'totale': 0.0
            }
        })
    
@app.route('/verifica_tappi_profilo', methods=['POST'])
def verifica_tappi_profilo():
    try:
        data = request.json
        profilo_id = data.get('profilo_id')

        profilo_id = profilo_id.split('_')[0]
        print(profilo_id)
        
        if not profilo_id:
            return jsonify({'success': False, 'message': 'Profilo ID mancante'})

        tappi_result = db.supabase.table('tappi')\
            .select('*')\
            .eq('prf_riferimento', profilo_id)\
            .execute()
        
        if tappi_result.data and len(tappi_result.data) > 0:
            tappi_disponibili = []
            for tappo in tappi_result.data:
                tappi_disponibili.append({
                    'id': tappo['codice'],
                    'codice': tappo['codice'],
                    'prezzo': float(tappo['prezzo']) if tappo['prezzo'] else 0.0,
                    'finitura': tappo['finitura'],
                    'lunghezza_interna': float(tappo['lunghezza_interna']) if tappo['lunghezza_interna'] else 0.0,
                    'lunghezza_esterna': float(tappo['lunghezza_esterna']) if tappo['lunghezza_esterna'] else 0.0,
                    'quantita': int(tappo['quantita']) if tappo['quantita'] else 1,
                    'forati': tappo['forati'],
                    'inclusi': tappo['inclusi']
                })
            
            return jsonify({
                'success': True,
                'has_tappi': True,
                'tappi_disponibili': tappi_disponibili
            })
        else:
            return jsonify({
                'success': True,
                'has_tappi': False
            })
            
    except Exception as e:
        logging.error(f"Errore in verifica_tappi_profilo: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/verifica_diffusori_profilo', methods=['POST'])
def verifica_diffusori_profilo():
    try:
        data = request.json
        profilo_id = data.get('profilo_id')

        profilo_id = profilo_id.split('_')[0]
        print(f"Verifica diffusori per profilo: {profilo_id}")

        if not profilo_id:
            return jsonify({'success': False, 'message': 'Profilo ID mancante'})

        diffusori_result = db.supabase.table('diffusori')\
            .select('*')\
            .eq('prf_riferimento', profilo_id)\
            .execute()

        if diffusori_result.data and len(diffusori_result.data) > 0:
            diffusore = diffusori_result.data[0]
            # Convert length from meters to millimeters
            lunghezza_metri = float(diffusore['lunghezza']) if diffusore['lunghezza'] else 0.0
            lunghezza_mm = lunghezza_metri * 1000

            diffusore_data = {
                'id': diffusore['codice'],
                'codice': diffusore['codice'],
                'prezzo': float(diffusore['prezzo']) if diffusore['prezzo'] else 0.0,
                'lunghezza': lunghezza_mm
            }

            return jsonify({
                'success': True,
                'has_diffusore': True,
                'diffusore': diffusore_data
            })
        else:
            return jsonify({
                'success': True,
                'has_diffusore': False
            })

    except Exception as e:
        logging.error(f"Errore in verifica_diffusori_profilo: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/verifica_staffe_profilo', methods=['POST'])
def verifica_staffe_profilo():
    try:
        data = request.json
        profilo_id = data.get('profilo_id')

        profilo_id = profilo_id.split('_')[0]
        print(f"Verifica staffe per profilo: {profilo_id}")

        if not profilo_id:
            return jsonify({'success': False, 'message': 'Profilo ID mancante'})

        staffe_result = db.supabase.table('staffe')\
            .select('*')\
            .eq('prf_riferimento', profilo_id)\
            .eq('incluso', False)\
            .execute()

        if staffe_result.data and len(staffe_result.data) > 0:
            staffa = staffe_result.data[0]

            staffa_data = {
                'id': staffa['codice'],
                'codice': staffa['codice'],
                'prezzo': float(staffa['prezzo']) if staffa['prezzo'] else 0.0,
                'incluso': staffa['incluso']
            }

            return jsonify({
                'success': True,
                'has_staffa': True,
                'staffa': staffa_data
            })
        else:
            return jsonify({
                'success': True,
                'has_staffa': False
            })

    except Exception as e:
        logging.error(f"Errore in verifica_staffe_profilo: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)